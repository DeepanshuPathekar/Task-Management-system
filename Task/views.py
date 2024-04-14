from django.shortcuts import render,redirect
from django.http import HttpResponse
from .models import Users, Tasks
import openpyxl

 
def home(request):
    return render(request, "index.html")

def export_data(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="user_task_data.xlsx"'
 
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "User Task Data"

    ws['A1'] = 'User Name'
    ws['B1'] = 'User Email'
    ws['C1'] = 'User Mobile'
    ws['D1'] = 'Task Detail'
    ws['E1'] = 'Task Type'

    users = Users.objects.all()
    for i, user in enumerate(users, start=2):
        ws[f'A{i}'] = user.name
        ws[f'B{i}'] = user.email
        ws[f'C{i}'] = user.mobile
        tasks = Tasks.objects.filter(user=user)
        for task in tasks:
            ws[f'D{i}'] = task.task_detail
            ws[f'E{i}'] = task.task_type
            i += 1

    wb.save(response)
    return response

def user_list(request):
    users = Users.objects.all()
    tasks = Tasks.objects.all()
    return render(request, "users.html", {"users": users, "tasks": tasks})

